"""
BuildQuote AI
Excel Export Service
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os


def export_excel(
    estimate,
    client_name="Customer",
    project_name="Construction Project"
):

        # ==========================================
    # CREATE WORKBOOK
    # ==========================================

    workbook = Workbook()

    sheet = workbook.active

    sheet.title = "Project"

    header_fill = PatternFill(
        fill_type="solid",
        start_color="1F4E78",
        end_color="1F4E78"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    center = Alignment(
        horizontal="center"
    )

    project = estimate.get(
        "project",
        {}
    )    # ==========================================
    # PROJECT INFORMATION
    # ==========================================

    rows = [

        ["Property", "Value"],

        ["Client", client_name],

        ["Project", project_name],

        ["County", project.get("County", "-")],

        ["Project Type", project.get("Project Type", "-")],

        ["House Type", project.get("House Type", "-")],

        ["Wall Material", project.get("Block Type", "-")],

        ["Roof Type", project.get("Roof Type", "-")],

        ["Bedrooms", project.get("Bedrooms", "-")],

        ["Length (m)", project.get("Length", 0)],

        ["Width (m)", project.get("Width", 0)],

        ["Wall Height (m)", project.get("Wall Height", 0)],

        ["Floor Area (m²)", project.get("Floor Area", 0)],

    ]

    for row in rows:

        sheet.append(row)

    # Style header

    for cell in sheet[1]:

        cell.fill = header_fill

        cell.font = header_font

        cell.alignment = center    # ==========================================
    # COST SUMMARY SHEET
    # ==========================================

    cost_sheet = workbook.create_sheet("Cost Summary")

    cost_sheet.append(["Item", "Amount (KES)"])

    cost_sheet.append([
        "Subtotal",
        estimate.get("subtotal", 0)
    ])

    cost_sheet.append([
        "VAT (16%)",
        estimate.get("vat", 0)
    ])

    cost_sheet.append([
        "Grand Total",
        estimate.get("grand_total", 0)
    ])

    # Header Style

    for cell in cost_sheet[1]:

        cell.fill = header_fill

        cell.font = header_font

        cell.alignment = center    # ==========================================
    # BOQ SHEET
    # ==========================================

    boq_sheet = workbook.create_sheet("BOQ")

    boq_sheet.append([
        "Section",
        "Item",
        "Value"
    ])

    boq = estimate.get(
        "boq",
        {}
    )

    if isinstance(boq, dict):

        for section, items in boq.items():

            if isinstance(items, dict):

                for item, value in items.items():

                    if isinstance(value, float):
                        value = round(value, 2)

                    boq_sheet.append([
                        section,
                        item,
                        value
                    ])

    # Style Header

    for cell in boq_sheet[1]:

        cell.fill = header_fill

        cell.font = header_font

        cell.alignment = center    # ==========================================
    # MATERIALS SHEET
    # ==========================================

    material_sheet = workbook.create_sheet("Materials")

    material_sheet.append([
        "Section",
        "Material",
        "Quantity"
    ])

    materials = estimate.get(
        "materials",
        {}
    )

    if isinstance(materials, dict):

        for section, data in materials.items():

            if isinstance(data, dict):

                for material, quantity in data.items():

                    if isinstance(quantity, float):
                        quantity = round(quantity, 2)

                    material_sheet.append([
                        section,
                        material,
                        quantity
                    ])

    # Style Header

    for cell in material_sheet[1]:

        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center    # ==========================================
    # LABOUR SHEET
    # ==========================================

    labour_sheet = workbook.create_sheet("Labour")

    labour_sheet.append([
        "Section",
        "Activity",
        "Cost (KES)"
    ])

    labour = estimate.get(
        "labour",
        {}
    )

    if isinstance(labour, dict):

        for section, data in labour.items():

            if isinstance(data, dict):

                for activity, cost in data.items():

                    if isinstance(cost, (int, float)):
                        cost = round(cost, 2)

                    labour_sheet.append([
                        section,
                        activity,
                        cost
                    ])

    # Style Header

    for cell in labour_sheet[1]:

        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center    # ==========================================
    # AUTO SIZE COLUMNS
    # ==========================================

    for ws in workbook.worksheets:

        for column_cells in ws.columns:

            length = 0

            column = get_column_letter(column_cells[0].column)

            for cell in column_cells:

                try:

                    length = max(
                        length,
                        len(str(cell.value))
                    )

                except Exception:
                    pass

            ws.column_dimensions[column].width = min(length + 4, 40)

    # ==========================================
    # SAVE WORKBOOK
    # ==========================================

    filename = (
        project_name.replace(" ", "_")
        + "_BuildQuote_Report.xlsx"
    )

    workbook.save(filename)

    return filename